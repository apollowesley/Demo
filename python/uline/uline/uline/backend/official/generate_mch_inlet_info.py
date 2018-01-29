# -*- coding: utf-8 -*-
from __future__ import division
from os import path, makedirs
import time
from openpyxl import Workbook
from collections import defaultdict
from uline.public import common, log

from uline.backend.__init__ import *

cur_dir = path.dirname(path.dirname(path.dirname(path.abspath(__file__))))
from uline.public.constants import PAYMENT, AUTH_STATUS, BALANCE_TYPE, BALANCE_WAY, DOWNLOAD_INFO_NUM_LIMIT, WX_USE_PARENT


# 任务队列
# 生成商户进件信息
@app.task
def generate_xls(user_id, dt_name, mch_name, dt_id, mch_id, create_at_start, create_at_end, activated_status,
                 auth_status, total_num):
    wb = Workbook()
    inlet_info_ws = wb.create_sheet(u'商户基本信息', 0)
    order_id = common.create_order_id()
    file_name, file_path, static_path = general_filename(user_id)
    if total_num > DOWNLOAD_INFO_NUM_LIMIT:
        gen_order_download_info(order_id, user_id, file_name)
    try:
        inlet_info_data = db_inlet_info(dt_name, mch_name, dt_id, mch_id, create_at_start, create_at_end,
                                        activated_status, auth_status)
        id_to_name = {data[0]: data[1] for data in inlet_info_data}

        industry_name = get_industry_name(id_to_name)
        gen_inlet_info(inlet_info_ws, inlet_info_data, industry_name)

        payment_info_data = db_mch_payment_info(id_to_name)
        payment_info_ws = wb.create_sheet(u'支付类型', 1)
        gen_mch_payment_info(payment_info_ws, payment_info_data)
        wb.save(file_path)
        if total_num > DOWNLOAD_INFO_NUM_LIMIT:
            modify_order_download_info(order_id, 2)
    except Exception as err:
        if total_num > DOWNLOAD_INFO_NUM_LIMIT:
            modify_order_download_info(order_id, 3)
        log.exception.info(err)
        return {'static_path': False}
    return {'static_path': static_path}


def gen_order_download_info(order_id, user_id, file_name):
    create_at = update_at = common.timestamp_now()
    query = """insert into order_download_info (order_id, user_id, file, type, status, platform, create_at, update_at)
            values (%s,%s,%s,%s,%s,%s,%s,%s)"""
    db.executeSQL(query, (order_id, user_id, file_name,
                          1, 1, 1, create_at, update_at))


def modify_order_download_info(order_id, status):
    update_at = common.timestamp_now()
    query = """update order_download_info set status=%s,update_at=%s where order_id=%s;"""
    db.executeSQL(query, (status, update_at, order_id))


def db_inlet_info(dt_name, mch_name, dt_id, mch_id, create_at_start, create_at_end, activated_status, auth_status):
    query = """
        SELECT s.* from (
            SELECT
              mch_inlet_info.mch_id,
              mch_inlet_info.mch_shortname,
              mch_inlet_info.mch_name,
              mch_user.wx_sub_mch_id,
              dt_user.dt_name,
              mch_inlet_info.province,
              mch_inlet_info.city,
              mch_inlet_info.address,
              mch_inlet_info.contact,
              mch_inlet_info.mobile,
              mch_inlet_info.email,
              mch_balance.balance_way,
              mch_balance.balance_type,
              mch_balance.balance_name,
              balance_bank_info.bank_name,
              balance_bank_info.bank_no,
              mch_balance.balance_account,
              mch_balance.id_card_no,
              mch_inlet_info.auth_status,
              mch_user.wx_use_parent,
              row_number()
              OVER (partition by mch_inlet_info.mch_id)
              as group_idx FROM mch_inlet_info
            INNER JOIN mch_user on mch_inlet_info.mch_id = mch_user.mch_id
              INNER JOIN dt_user on dt_user.dt_id = mch_inlet_info.dt_id
              INNER JOIN mch_balance ON mch_inlet_info.mch_id = mch_balance.mch_id
              INNER JOIN balance_bank_info on mch_balance.bank_no = balance_bank_info.bank_no
            WHERE (mch_inlet_info.activated_status=%(activated_status)s OR %(activated_status)s IS NULL )
              AND (mch_inlet_info.auth_status=%(auth_status)s OR %(auth_status)s IS NULL )
              AND (dt_user.dt_name ~ %(dt_name)s or %(dt_name)s IS NULL  )
              and (mch_inlet_info.cs_id is null)
              AND (dt_user.dt_id::VARCHAR ~ %(dt_id)s::VARCHAR or %(dt_id)s::VARCHAR IS NULL  )
              AND (mch_user.mch_id::VARCHAR ~ %(mch_id)s::VARCHAR or %(mch_id)s::VARCHAR IS NULL  )
              AND (mch_inlet_info.mch_name ~ %(mch_name)s OR %(mch_name)s IS NULL )
              AND (mch_inlet_info.create_at BETWEEN %(create_at_start)s::TIMESTAMP AND %(create_at_end)s::TIMESTAMP
              OR %(create_at_start)s IS NULL OR %(create_at_end)s IS NULL )
            ORDER BY mch_inlet_info.create_at) s
                  where s.group_idx=1
                """

    ret = db.selectSQL(query, {
        'activated_status': activated_status,
        'auth_status': auth_status,
        'mch_name': mch_name,
        'mch_id': mch_id,
        'dt_id': dt_id,
        'dt_name': dt_name,
        'create_at_start': create_at_start,
        'create_at_end': create_at_end,
    }, fetchone=False)
    return ret


def get_industry_name(mch_id_to_name):
    query = """select
                mch_inlet_info.mch_id,
                industry_uline_info.industry_name
                from mch_inlet_info
                inner join industry_uline_info on industry_uline_info.industry_code = mch_inlet_info.u_ind_code
                where mch_inlet_info.mch_id = ANY(%s);"""
    used_ret = db.selectSQL(query, (list(mch_id_to_name),), fetchone=False)

    query = """select
                mch_inlet_info.mch_id,
                industry_uline_info.industry_name
                from mch_inlet_info
                inner join industry_uline_info on industry_uline_info.wx_ind_code = mch_inlet_info.old_ind_code
                where mch_inlet_info.mch_id = ANY(%s);"""
    unused_ret = db.selectSQL(query, (list(mch_id_to_name),), fetchone=False)

    ret = list(used_ret)
    unused_ret = list(unused_ret)
    ret.extend(unused_ret)
    return ret


def db_mch_payment_info(mch_id_to_name):
    """
    :param mch_id_to_name: id到name的映射
    :return:
    """
    # Furthermore ANY can also work with empty lists, whereas IN () is a SQL
    # syntax error.
    if not mch_id_to_name:
        return
    query = """SELECT mch_id,payment_type,payment_rate,activated_status
            FROM mch_payment WHERE mch_id = ANY(%s);
            """

    ret = db.selectSQL(query, (list(mch_id_to_name),), fetchone=False)
    return ret


def gen_inlet_info(ws, datas, industry_name):
    fields = [u'商户编号', u'商户简称', u'商户名称', u'微信商户编号', u'渠道商名称', u'省份', u'城市', u'联系地址', u'联系人', u'联系电话',
              u'电子邮箱', u'结算方式', u'结算账号类型', u'结算户名', u'支行名称', u'支行联行号', u'结算账号', u'身份证号码',
              u'审核状态', u'是否使用渠道商微信appid', u'行业类别']
    # 渠道基本信息表
    ws.append(fields)

    industry_names = {}
    for dt_id, name in industry_name:
        industry_names[dt_id] = name

    for data in datas:
        data = list(data)
        data[11], data[12], data[18], data[19] = BALANCE_WAY[str(data[11])], BALANCE_TYPE[str(data[12])], AUTH_STATUS[
            str(data[18])], WX_USE_PARENT[str(data[19])]
        data.extend([industry_names[data[0]]])
        del data[-2]
        ws.append(data)


def gen_mch_payment_info(ws, datas):
    # 支付类型表
    fields = [u'商户名称', u'支付类型', u'结算费率(‰)', u'激活状态']
    ws.append(fields)

    for data in datas:
        data = list(data)
        data[1], data[2], data[3] = PAYMENT[str(
            data[1])], data[2] / 10, AUTH_STATUS[str(data[3])]
        ws.append(data)


def general_filename(user_id):
    _ts = str(time.time())
    filename = "mch_inlet_export_" + _ts + ".xlsx"
    user_dl_path = path.join(cur_dir, "static/downloads/", str(user_id))
    if not path.exists(user_dl_path):
        makedirs(user_dl_path)
    file_path = path.join(user_dl_path, filename)
    static_path = path.join("/static/downloads/", str(user_id), filename)
    return filename, file_path, static_path
