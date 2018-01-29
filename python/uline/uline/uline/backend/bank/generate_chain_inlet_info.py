# -*- coding: utf-8 -*-
from __future__ import division
from os import path, makedirs
import time
from openpyxl import Workbook
from collections import defaultdict

from uline.backend.__init__ import *

cur_dir = path.dirname(path.dirname(path.dirname(path.abspath(__file__))))
from uline.public.constants import PAYMENT, AUTH_STATUS, BALANCE_TYPE, BALANCE_WAY, DOWNLOAD_INFO_NUM_LIMIT
from uline.public import common, log


# 任务队列
@app.task
def generate_xls(user_id, dt_name, dt_id, create_at_start, create_at_end, activated_status, auth_status, total_num, parent_id, parent_name):
    """
    渠道编号	渠道名称	行业类别   	省份    	城市 	联系地址	联系人	联系电话	电子邮箱	结算方式	结算账号类型	结算户名	结算银行	支行名称	支行联行号 结算账号	身份证号码	审核状态
    商户编号	商户名称	支付类型 	结算费率（‰）	激活状态
    :return:
    """
    wb = Workbook()
    inlet_info_ws = wb.create_sheet(u'渠道基本信息', 0)

    order_id = common.create_order_id()
    file_name, file_path, static_path = general_filename(user_id)
    if total_num > DOWNLOAD_INFO_NUM_LIMIT:
        gen_order_download_info(order_id, user_id, file_name)
    try:
        inlet_info_data = db_inlet_info(dt_name, dt_id, create_at_start,
                                        create_at_end, activated_status, auth_status, parent_id, parent_name)
        id_to_name = {data[0]: data[1] for data in inlet_info_data}

        industry_name = get_industry_name(id_to_name)
        gen_inlet_info(inlet_info_ws, inlet_info_data, industry_name)

        payment_info_data = db_mch_payment_info(id_to_name)
        payment_info_ws = wb.create_sheet(u'支付类型', 1)
        gen_mch_payment_info(payment_info_ws, payment_info_data)

        wb.save(file_path)

        if total_num > DOWNLOAD_INFO_NUM_LIMIT:
            modify_order_download_info(order_id, 2)

    except Exception as err:
        if total_num > DOWNLOAD_INFO_NUM_LIMIT:
            modify_order_download_info(order_id, 3)
        log.exception.info(err)
        return {'static_path': False}
    return {'static_path': static_path}


def gen_order_download_info(order_id, user_id, file_name):
    create_at = update_at = common.timestamp_now()
    query = """insert into order_download_info (order_id, user_id, file, type, status, platform, create_at, update_at)
            values (%s,%s,%s,%s,%s,%s,%s,%s)"""
    db.executeSQL(query, (order_id, user_id, file_name,
                          2, 1, 1, create_at, update_at))


def modify_order_download_info(order_id, status):
    update_at = common.timestamp_now()
    query = """update order_download_info set status=%s,update_at=%s where order_id=%s;"""
    db.executeSQL(query, (status, update_at, order_id))


def db_inlet_info(dt_name, dt_id, create_at_start, create_at_end, activated_status, auth_status, parent_id, parent_name):
    query = """
    SELECT s.* from (
        select
        dt_inlet_info.dt_id,
        dt_inlet_info.dt_name,
        dt_inlet_info.province,
        dt_inlet_info.city,
        dt_inlet_info.address,
        dt_inlet_info.contact,
        dt_inlet_info.mobile,
        dt_inlet_info.email,
        dt_balance.balance_way,
        dt_balance.balance_type,
        balance_bank_info.bank_name,
        balance_bank_info.bank_no,
        dt_balance.balance_account,
        dt_balance.id_card_no,
        dt_inlet_info.auth_status,
        dt_inlet_info.dt_sub_id,
        dt_subuser.dt_sub_name,
      row_number()
      OVER (partition by dt_inlet_info.dt_id)
      as group_idx FROM dt_inlet_info
        inner join dt_user on dt_user.dt_id = dt_inlet_info.parent_id
        inner join dt_balance on dt_balance.dt_id = dt_inlet_info.dt_id
        inner join balance_bank_info on dt_balance.bank_no = balance_bank_info.bank_no
        left join dt_subuser on dt_inlet_info.dt_sub_id = dt_subuser.dt_sub_id
        where (dt_inlet_info.activated_status = %(activated_status)s or %(activated_status)s is null)
        and (dt_inlet_info.auth_status = %(auth_status)s or %(auth_status)s is null)
        and (dt_inlet_info.dt_name ~ %(dt_name)s or %(dt_name)s is null)
        and (dt_user.dt_id=%(dt_boss_id)s or %(dt_boss_id)s is null)
        and (dt_inlet_info.parent_id is not null)
        and (dt_user.dt_name=%(dt_boss_name)s or %(dt_boss_name)s is null)
        and (dt_inlet_info.dt_id::VARCHAR ~ %(dt_id)s::VARCHAR or %(dt_id)s::VARCHAR is null)
        and (dt_inlet_info.create_at BETWEEN %(create_at_start)s::TIMESTAMP
         and %(create_at_end)s::TIMESTAMP OR %(create_at_start)s IS NULL OR %(create_at_end)s IS NULL )
        ORDER BY dt_inlet_info.create_at) s
                  where s.group_idx=1
        """

    ret = db.selectSQL(query, {
        'activated_status': activated_status,
        'auth_status': auth_status,
        'dt_name': dt_name,
        'dt_boss_name': parent_name,
        'dt_id': dt_id,
        'dt_boss_id': parent_id,
        'create_at_start': create_at_start,
        'create_at_end': create_at_end,
    }, fetchone=False)
    return ret


def db_mch_payment_info(dt_id_to_name):
    if not dt_id_to_name:
        return
    query = """SELECT dt_id,payment_type,payment_rate,activated_status
            FROM dt_payment WHERE dt_id = ANY(%s);
            """

    ret = db.selectSQL(query, (list(dt_id_to_name),), fetchone=False)
    return ret


def get_wx_ali_ind_code(dt_id_to_name):
    query = """select
            dt_inlet_info.dt_id,
            industry_ali_info.industry_name,
            industry_uline_info.industry_name
            from dt_inlet_info
            inner join industry_uline_info on industry_uline_info.industry_code = dt_inlet_info.u_ind_code
            inner join industry_ali_info on industry_ali_info.industry_code = dt_inlet_info.ali_ind_code
            where dt_inlet_info.dt_id = ANY(%s);"""
    ret = db.selectSQL(query, (list(dt_id_to_name),), fetchone=False)
    return ret


def get_industry_name(dt_id_to_name):
    query = """select
            dt_inlet_info.dt_id,
            industry_uline_info.industry_name
            from dt_inlet_info
            inner join industry_uline_info on industry_uline_info.industry_code = dt_inlet_info.u_ind_code
            where dt_inlet_info.dt_id = ANY(%s);"""
    used_ret = db.selectSQL(query, (list(dt_id_to_name),), fetchone=False)

    query = """select
            dt_inlet_info.dt_id,
            industry_uline_info.industry_name
            from dt_inlet_info
            inner join industry_uline_info on industry_uline_info.wx_ind_code = dt_inlet_info.old_ind_code
            where dt_inlet_info.dt_id = ANY(%s);"""
    unused_ret = db.selectSQL(query, (list(dt_id_to_name),), fetchone=False)

    ret = list(used_ret)
    unused_ret = list(unused_ret)
    ret.extend(unused_ret)
    return ret


def gen_inlet_info(ws, datas, industry_name):
    fields = [u'连锁商户编号', u'连锁商户名称', u'省份', u'城市', u'联系地址', u'联系人', u'联系电话', u'电子邮箱',
              u'结算方式', u'结算账号类型', u'支行名称', u'支行联行号', u'结算账号', u'身份证号码', u'审核状态', u'业务员编号', u'业务员名称',
              u'行业类别']
    # 渠道基本信息表
    ws.append(fields)

    industry_names = {}
    for dt_id, name in industry_name:
        industry_names[dt_id] = name

    for data in datas:
        data = list(data)
        data[8], data[9], data[14] = BALANCE_WAY[str(data[8])], BALANCE_TYPE[str(data[9])], AUTH_STATUS[
            str(data[14])]
        data.extend([industry_names[data[0]]])
        del data[-2]
        ws.append(data)


def gen_mch_payment_info(ws, datas):
    # 支付类型表
    fields = [u'渠道商名称', u'支付类型', u'结算费率(‰)', u'激活状态']
    ws.append(fields)

    for data in datas:
        data = list(data)
        data[1], data[2], data[3] = PAYMENT[str(
            data[1])], data[2] / 10, AUTH_STATUS[str(data[3])]
        ws.append(data)


def general_filename(user_id):
    _ts = str(time.time())
    filename = "dt_inlet_export_" + _ts + ".xlsx"
    user_dl_path = path.join(cur_dir, "static/downloads/", str(user_id))
    if not path.exists(user_dl_path):
        makedirs(user_dl_path)
    file_path = path.join(user_dl_path, filename)
    static_path = path.join("/static/downloads/", str(user_id), filename)
    return filename, file_path, static_path
